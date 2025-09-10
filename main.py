import os
import re
import traceback
import unicodedata
from datetime import datetime
from flask import Flask, request, jsonify, render_template, send_file
from werkzeug.utils import secure_filename
from dotenv import load_dotenv
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from email.message import EmailMessage
import smtplib
from email_validator import validate_email, EmailNotValidError

load_dotenv()

BASE_XLSX = os.getenv("BASE_XLSX", "baza.xlsx")
SMTP_HOST = os.getenv("SMTP_HOST")
SMTP_PORT = int(os.getenv("SMTP_PORT", 587)) if os.getenv("SMTP_PORT") else 587
SMTP_USER = os.getenv("SMTP_USER")
SMTP_PASS = os.getenv("SMTP_PASS")
EMAIL_FROM = os.getenv("EMAIL_FROM", SMTP_USER)
EMAIL_FROM_NAME = os.getenv("EMAIL_FROM_NAME", "Formatki OBI")
LOGO_URL = os.getenv("LOGO_URL", "")

TMP_DIR = os.path.join(os.getcwd(), "tmp")
os.makedirs(TMP_DIR, exist_ok=True)

app = Flask(__name__, static_folder="static", template_folder="templates")
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

ALLOWED_DOMAIN = "obi.pl"


def _load_df():
    if not os.path.exists(BASE_XLSX):
        raise FileNotFoundError(f"Plik nie znaleziony: {BASE_XLSX}")
    df = pd.read_excel(BASE_XLSX, sheet_name="Arkusz1", header=0, dtype=str)
    df = df.fillna("")

    def _norm(v):
        if isinstance(v, str):
            s = v.strip()
            if (s.startswith("'") and s.endswith("'")) or (s.startswith('"') and s.endswith('"')):
                s = s[1:-1].strip()
            return s
        return v

    df = df.map(_norm)
    return df


def _detect_columns(df):
    cols = {c.strip().lower(): c for c in df.columns}
    if 'gt' in cols and 'kw' in cols and 'pion' in cols:
        return cols['gt'], cols['kw'], cols['pion']
    col0 = df.columns[0]
    col1 = df.columns[1] if len(df.columns) > 1 else df.columns[0]
    col2 = df.columns[2] if len(df.columns) > 2 else df.columns[0]
    return col0, col1, col2


def _safe_sheet_name(name, existing_names=None):
    if existing_names is None:
        existing_names = set()
    if not name:
        base = "Sheet"
    else:
        base = re.sub(r'[:\\\/\?\*\[\]]', '-', str(name)).strip()
    if not base:
        base = "Sheet"
    max_len = 31
    base = base[:max_len]
    candidate = base
    i = 1
    while candidate in existing_names:
        suffix = f"_{i}"
        allowed = max_len - len(suffix)
        candidate = (base[:allowed] + suffix) if allowed > 0 else base[:max_len]
        i += 1
    existing_names.add(candidate)
    return candidate


def _cmp_norm_for_match(s):
    if s is None:
        return ""
    s = str(s)
    s = s.strip()
    s = unicodedata.normalize('NFKC', s)
    s = re.sub(r'[\u2715\u00D7\u2716\u2717\u2718]', '', s)
    s = re.sub(r'\s+', ' ', s)
    return s.lower()


def _compress_row_values_left(ws, row_idx, col_start_idx, col_end_idx):
    vals = []
    for c in range(col_start_idx, col_end_idx + 1):
        v = ws.cell(row=row_idx, column=c).value
        if v is not None and str(v).strip() != "":
            vals.append(v)
    for c in range(col_start_idx, col_end_idx + 1):
        ws.cell(row=row_idx, column=c).value = None
    for i, v in enumerate(vals):
        ws.cell(row=row_idx, column=col_start_idx + i).value = v


def _style_workbook(path):
    from openpyxl.utils import get_column_letter
    wb = load_workbook(path)
    header_fill = PatternFill(start_color="F47B20", end_color="F47B20", fill_type="solid")
    header_font = Font(bold=True)
    header_row_height = 30
    for name in wb.sheetnames:
        if name == "Wymagania":
            ws = wb[name]
            try:
                ws.column_dimensions['A'].width = 40
            except Exception:
                pass
            try:
                for r in (10, 11, 12):
                    cell = ws.cell(row=r, column=1)
                    if cell.value:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                        lines = str(cell.value).count("\n") + 1
                        try:
                            ws.row_dimensions[r].height = max(20, 15 * lines)
                        except Exception:
                            pass
            except Exception:
                pass
            continue
        ws = wb[name]
        try:
            first_cell = ws.cell(row=1, column=1).value
            if (first_cell is None or str(first_cell).strip() == "") and ws.max_column > 1:
                ws.delete_cols(1)
        except Exception:
            pass
        if ws.max_row >= 1:
            for cell in list(ws[1]):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        try:
            ws.row_dimensions[1].height = header_row_height
        except Exception:
            pass
        try:
            _compress_row_values_left(ws, 2, 9, min(23, ws.max_column))
        except Exception:
            app.logger.debug("compress_row_values_left failed for sheet %s", name)
        try:
            max_rows_to_check = min(ws.max_row, 20)
            for col_idx in range(1, ws.max_column + 1):
                max_len = 0
                for row_idx in range(1, max_rows_to_check + 1):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if val is None:
                        continue
                    s = str(val).replace("\n", " ")
                    if len(s) > max_len:
                        max_len = len(s)
                if max_len <= 0:
                    width = 8
                else:
                    width = min(max(10, int(max_len * 1.1)), 60)
                col_letter = get_column_letter(col_idx)
                try:
                    ws.column_dimensions[col_letter].width = width
                except Exception:
                    pass
        except Exception:
            app.logger.debug("auto column width failed for sheet %s", name)
        try:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and "\n" in cell.value:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
        except Exception:
            pass
    wb.save(path)
    wb.close()


def _extract_attribute_from_row(row, desired_attributes, punktor_cols):
    import re
    def _clean_val(v):
        if v is None:
            return ""
        if not isinstance(v, str):
            v = str(v)
        s = v.strip()
        if (s.startswith("'") and s.endswith("'")) or (s.startswith('"') and s.endswith('"')):
            s = s[1:-1].strip()
        return s
    out = {}
    cols_map = {c.strip().lower(): c for c in row.index}
    for attr in desired_attributes:
        found = ""
        if attr.strip().lower() in cols_map:
            found = _clean_val(row[cols_map[attr.strip().lower()]])
            out[attr] = found
            continue
        label_clean = re.sub(r"[:\s]+$", "", attr.strip().lower())
        search_cols = list(punktor_cols) + [c for c in row.index if c not in punktor_cols]
        for c in search_cols:
            v = row.get(c, "")
            if v is None:
                continue
            v_clean = _clean_val(v)
            low = v_clean.lower()
            if low.startswith(label_clean):
                parts = re.split(r'[:=\-]', v_clean, maxsplit=1)
                if len(parts) >= 2:
                    found = parts[1].strip()
                else:
                    found = v_clean[len(label_clean):].strip()
                break
            if label_clean in low:
                parts = re.split(re.escape(label_clean), low, maxsplit=1)
                if len(parts) >= 2 and parts[1].strip():
                    found = parts[1].strip()
                    break
        if not found:
            try:
                idx = desired_attributes.index(attr)
                if idx < len(punktor_cols):
                    found = _clean_val(row.get(punktor_cols[idx], ""))
            except ValueError:
                pass
        out[attr] = found
    return out


def _wrap_every_n_words(s, n=5):
    if not s:
        return s
    parts = re.split(r'(\s+)', s)
    words = []
    for p in parts:
        if p.strip() == "":
            continue
        words.append(p)
    lines = []
    for i in range(0, len(words), n):
        lines.append(" ".join(words[i:i+n]))
    return "\n".join(lines)


def _write_excel_and_format(pion, gt_list, kw_list, df, desired_base, desired_attributes, filename):
    tmp_path = os.path.join(TMP_DIR, secure_filename(filename))
    found_any = False
    used_sheet_names = set()
    gt_col, kw_col, pion_col = _detect_columns(df)
    app.logger.info("Detected columns: GT=%s, KW=%s, PION=%s", gt_col, kw_col, pion_col)
    punktor_cols = [c for c in df.columns if str(c).strip().lower().startswith("punktor")]
    if not punktor_cols:
        candidate_idxs = list(range(10, min(len(df.columns), 30)))
        punktor_cols = [df.columns[i] for i in candidate_idxs if i < len(df.columns)]
    app.logger.info("Punktor cols sample: %s", punktor_cols[:8])
    def _clean_val(v):
        if v is None:
            return ""
        if not isinstance(v, str):
            v = str(v)
        s = v.strip()
        if (s.startswith("'") and s.endswith("'")) or (s.startswith('"') and s.endswith('"')):
            s = s[1:-1].strip()
        return s
    with pd.ExcelWriter(tmp_path, engine="xlsxwriter") as writer:
        for gt in gt_list:
            for kw in kw_list:
                sel = df[
                    (df[pion_col].astype(str).str.strip().str.lower() == str(pion).strip().lower())
                    & (df[gt_col].astype(str).str.strip().str.lower() == str(gt).strip().lower())
                    & (df[kw_col].astype(str).str.strip().str.lower() == str(kw).strip().lower())
                ]
                app.logger.info("Filter result for GT=%s KW=%s: rows=%d", gt, kw, len(sel))
                if sel.shape[0] == 0:
                    continue
                found_any = True
                first_row = sel.iloc[0]
                dyn_headers = []
                for pc in punktor_cols:
                    val = _clean_val(first_row.get(pc, ""))
                    if val:
                        if not val.endswith(":") and not val.endswith(":"):
                            pass
                        if val not in dyn_headers:
                            dyn_headers.append(val)
                if not dyn_headers and desired_attributes:
                    dyn_headers = desired_attributes.copy()
                all_columns = desired_base + dyn_headers
                rows_out = []
                for _, row in sel.iterrows():
                    out_row = {}
                    for base_col in desired_base:
                        matched = None
                        for c in row.index:
                            if str(c).strip().lower() == base_col.strip().lower():
                                matched = c
                                break
                        if matched is not None:
                            val = row.get(matched, "")
                            out_row[base_col] = "" if pd.isna(val) else val
                        else:
                            out_row[base_col] = ""
                    for h in dyn_headers:
                        out_row[h] = ""
                    rows_out.append(out_row)
                out_df = pd.DataFrame(rows_out, columns=all_columns)
                raw_name = f"{kw}"
                sheet_name = _safe_sheet_name(raw_name, existing_names=used_sheet_names)
                try:
                    out_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    app.logger.info("Wrote sheet: %s rows=%d headers=%s", sheet_name, len(out_df), all_columns)
                except Exception as ex:
                    app.logger.exception("Error writing sheet %s: %s", sheet_name, str(ex))
        if str(pion).strip().lower() == "oświetlenie" and (not gt_list):
            sel = df[df[pion_col].astype(str).str.strip().str.lower() == "oświetlenie"]
            if sel.shape[0] > 0:
                drop_cols = [c for c in ["GT", "KW", "PION", "Podział"] if c in sel.columns]
                sel2 = sel.drop(columns=drop_cols, errors="ignore")
                sheet_name = _safe_sheet_name("Oświetlenie", existing_names=used_sheet_names)
                sel2.to_excel(writer, sheet_name=sheet_name, index=False)
                app.logger.info("Wrote Oświetlenie sheet %s rows=%d", sheet_name, sel2.shape[0])
                found_any = True
        reqs = [
            '📸 Wymagania dotyczące zdjęć:',
            '- Zdjęcia minimum 1500 px na krótszy bok',
            '- Format .JPG',
            '- Packshot',
            '- Aranżacyjne',
            '- Więcej zdjęć = lepiej',
            '- Opisane numerem OBI lub EAN',
            'Wymagania dotyczące opisu i tytułu:',
            '- Tytuł artykułu online ma limit do 80 znaków',
            '- Opis artykułu powinien zawierać najważniejsze informacje opisowe z limitem 3997 znaków (3515 bez spacji), prosimy o podanie opisu artykułu z uwzględnieniem najważniejszych cech/zalet/zastosowań. Celem jest zebranie wszystkich ważnych informacji.',
            '- Dane znajdujące się w nawiasach klamrowych („{}”) stanowią możliwe opcje do wyboru — należy wybrać jedną z nich i wpisać ją w komórkę poniżej',
            '- Dane producenta - GPSR, są to dane, które pokazują się na stronie obi.pl jako dane wytwórcy, dane jakie należy podać to: Pełna nazwa firmy, adres siedziby oraz adres e-mail'
        ]
        processed = []
        for i, line in enumerate(reqs):
            if i in (9, 10, 11):
                processed.append(_wrap_every_n_words(line, 5))
            else:
                processed.append(line)
        pd.DataFrame(processed).to_excel(writer, sheet_name="Wymagania", index=False, header=False)
    app.logger.info("_write_excel_and_format finished; found_any=%s tmp_path=%s", found_any, tmp_path)
    try:
        _style_workbook(tmp_path)
    except Exception:
        app.logger.exception("Error styling workbook %s", tmp_path)
    return tmp_path, found_any


def _create_excel_for_selection(pion, gt_list, kw_list):
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    filename = f"Formatki-{pion}-{timestamp}.xlsx"
    df = _load_df()
    desired_base = [
        "EAN",
        "Nr. Art dostawcy",
        "Gwarancja: (lata)",
        "Opis artykułu (3997 znaków (3515 bez spacji))",
        "W zestawie:",
        "Dane producenta - GPSR:"
    ]
    desired_attributes = [
        "Moc [W]:",
        "Liczba biegów:",
        "Maks. prędkość obrotowa [obr/min]:",
        "Mocowanie mieszadła:",
        "Maksymalna średnica mieszadła [mm]:",
        "Gwarancja: {jeśli powyżej 2 lat}"
    ]
    tmp_path, found_any = _write_excel_and_format(pion, gt_list, kw_list, df, desired_base, desired_attributes, filename)
    return tmp_path, filename, found_any


def _send_email_with_attachment(to_emails, subject, html_body, attachment_path):
    if not SMTP_HOST or not SMTP_USER or not SMTP_PASS:
        raise RuntimeError("SMTP nie jest skonfigurowany (SMTP_HOST/SMTP_USER/SMTP_PASS).")
    valid = []
    for e in to_emails:
        try:
            v = validate_email(e)
            valid.append(v["email"])
        except EmailNotValidError:
            app.logger.warning("Invalid email skipped: %s", e)
    if not valid:
        raise ValueError("Brak poprawnych adresów e-mail do wysłania.")
    msg = EmailMessage()
    msg["From"] = f"{EMAIL_FROM_NAME} <{EMAIL_FROM or SMTP_USER}>"
    msg["To"] = ", ".join(valid)
    msg["Subject"] = subject
    msg.set_content("Wiadomość w HTML. Jeśli nie widzisz treści, otwórz e-mail w formacie HTML.")
    msg.add_alternative(html_body, subtype="html")
    with open(attachment_path, "rb") as f:
        data = f.read()
    maintype = "application"
    subtype = "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    msg.add_attachment(data, maintype=maintype, subtype=subtype, filename=os.path.basename(attachment_path))
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
        s.starttls()
        s.login(SMTP_USER, SMTP_PASS)
        s.send_message(msg)
    return True

@app.route("/")
@app.route("/index")
def index2():
    return render_template("index.html")

@app.route("/api/get_data_structure", methods=["GET"])
def api_get_data_structure():
    try:
        df = _load_df()
        gt_col, kw_col, pion_col = _detect_columns(df)
        structure = {}
        for _, row in df.iterrows():
            gt = str(row[gt_col]).strip()
            kw = str(row[kw_col]).strip()
            pion = str(row[pion_col]).strip()
            if not (gt and kw and pion):
                continue
            structure.setdefault(pion, {})
            structure[pion].setdefault(gt, set()).add(kw)
        out = {p: {g: sorted(list(kws)) for g, kws in mp.items()} for p, mp in structure.items()}
        return jsonify(out)
    except Exception as e:
        app.logger.exception("get_data_structure error")
        return jsonify({"error": str(e)}), 500

@app.route("/api/get_gt", methods=["GET"])
def api_get_gt():
    pion = request.args.get("pion", "")
    try:
        df = _load_df()
        gt_col, _, pion_col = _detect_columns(df)
        sel = df[df[pion_col].astype(str).str.strip().str.lower() == str(pion).strip().lower()]
        gts = sorted(sel[gt_col].astype(str).str.strip().unique())
        return jsonify(list(gts))
    except Exception as e:
        app.logger.exception("get_gt error")
        return jsonify({"error": str(e)}), 500

@app.route("/api/get_kw_for_gt_list", methods=["POST"])
def api_get_kw_for_gt_list():
    data = request.get_json(force=True)
    gt_list = data.get("gtList", []) or []
    try:
        df = _load_df()
        gt_col, kw_col, pion_col = _detect_columns(df)
        kws = set()
        for gt in gt_list:
            if not gt:
                continue
            matches = df[df[gt_col].astype(str).str.strip().str.lower() == str(gt).strip().lower()]
            for v in matches[kw_col].astype(str).tolist():
                if v and str(v).strip():
                    kws.add(str(v).strip())
        return jsonify(sorted(kws))
    except Exception as e:
        app.logger.exception("get_kw_for_gt_list error")
        return jsonify({"error": str(e)}), 500

@app.route("/api/resolve_gt_codes", methods=["POST"])
def api_resolve_gt_codes():
    data = request.get_json(force=True)
    pion = data.get("pion", "")
    raw = data.get("raw", "")
    codes = []
    if isinstance(raw, str):
        codes = [s.strip() for s in raw.split(",") if s.strip()]
    try:
        df = _load_df()
        gt_col, kw_col, pion_col = _detect_columns(df)
        dfp = df[df[pion_col].astype(str).str.strip().str.lower() == str(pion).strip().lower()]
        full = set()
        for code in codes:
            for val in dfp[gt_col].astype(str).tolist():
                if str(val).lower().startswith(code.lower()):
                    full.add(str(val).strip())
        return jsonify(sorted(full))
    except Exception as e:
        app.logger.exception("resolve_gt_codes error")
        return jsonify({"error": str(e)}), 500

@app.route("/_debug_rows", methods=["POST"])
def _debug_rows():
    data = request.get_json(force=True)
    pion = data.get("pion", "")
    gt = data.get("gt", "")
    kw = data.get("kw", "")
    try:
        df = _load_df()
        gt_col, kw_col, pion_col = _detect_columns(df)
        sel = df[
            (df[pion_col].astype(str).str.strip().str.lower() == str(pion).strip().lower())
            & (df[gt_col].astype(str).str.strip().str.lower() == str(gt).strip().lower())
            & (df[kw_col].astype(str).str.strip().str.lower() == str(kw).strip().lower())
        ]
        sample = sel.head(40).fillna("").to_dict(orient="records")
        return jsonify({"count": int(sel.shape[0]), "sample": sample})
    except Exception as e:
        app.logger.exception("_debug_rows error")
        return jsonify({"error": str(e)}), 500

@app.route("/api/generate_debug", methods=["POST"])
def api_generate_debug():
    try:
        data = request.get_json(force=True)
        pion = data.get("pion", "").strip()
        gt_list = data.get("gtList", []) or []
        kw_list = data.get("kwList", []) or []
        if not pion or (pion.lower() != "oświetlenie" and (not gt_list or not kw_list)):
            return jsonify({"success": False, "error": "Brakuje parametrów (pion/gtList/kwList)."}), 400
        tmp_path, filename, found_any = _create_excel_for_selection(pion, gt_list, kw_list)
        if not os.path.exists(tmp_path):
            return jsonify({"success": False, "error": "Plik nie został utworzony."}), 500
        return send_file(tmp_path, as_attachment=True, download_name=filename)
    except Exception as e:
        tb = traceback.format_exc()
        app.logger.error("Exception in api_generate_debug:\n%s", tb)
        return jsonify({"success": False, "error": str(e), "traceback": tb}), 500

@app.route("/api/generate", methods=["POST"])
def api_generate():
    try:
        data = request.get_json(force=True)
        pion = data.get("pion", "").strip()
        gt_list = data.get("gtList", []) or []
        kw_list = data.get("kwList", []) or []
        emails_raw = data.get("emails", data.get("email", "")) or ""
        if isinstance(emails_raw, str):
            emails_input = [e.strip() for e in emails_raw.split(",") if e.strip()]
        elif isinstance(emails_raw, list):
            emails_input = [e.strip() for e in emails_raw if e and str(e).strip()]
        else:
            emails_input = []
        emails = []
        for e in emails_input:
            try:
                v = validate_email(e)
                addr = v["email"]
                if addr.lower().endswith("@" + ALLOWED_DOMAIN):
                    emails.append(addr)
                else:
                    app.logger.info("Skipping non-allowed domain: %s", addr)
            except EmailNotValidError:
                app.logger.warning("Invalid email skipped: %s", e)
        if not pion or (pion.lower() != "oświetlenie" and (not gt_list or not kw_list)) or not emails:
            return jsonify({"success": False, "error": f"Brakuje parametrów (pion/gtList/kwList) lub brak poprawnych adresów z domeny @{ALLOWED_DOMAIN}."}), 400
        tmp_path, filename, found_any = _create_excel_for_selection(pion, gt_list, kw_list)
        if not os.path.exists(tmp_path):
            return jsonify({"success": False, "error": "Plik nie został utworzony."}), 500
        logo_html = f'<img src="{LOGO_URL}" alt="Logo" style="max-height:40px; margin-bottom:8px;" />' if LOGO_URL else ""
        bg = "#F47B20"
        html_body = f"""
        <html>
        <body style="font-family:Arial, sans-serif; background:{bg}; color:#ffffff; padding:20px;">
            <div style="max-width:680px; margin:0 auto; background:#ffffff; color:#000; padding:20px; border-radius:8px;">
                <div style="display:flex; align-items:center; gap:12px;">
                    {logo_html}
                    <h2 style="color:{bg}; margin:0;">Twoje formatki</h2>
                </div>
                <p>Cześć,<br>W załączeniu znajdziesz wygenerowany plik z formatkami dla pionu <strong>{pion}</strong>.</p>
                <p>Pozdrawiamy,<br>Zespół Product Content</p>
            </div>
        </body>
        </html>
        """
        _send_email_with_attachment(emails, f"Twój plik z formatkami - {pion}", html_body, tmp_path)
        return jsonify({"success": True, "message": "Wysłano e-maile."})
    except Exception as e:
        tb = traceback.format_exc()
        app.logger.error("Exception in api_generate:\n%s", tb)
        return jsonify({"success": False, "error": str(e), "traceback": tb}), 500

if __name__ == "__main__":
    app.run(debug=False, host="0.0.0.0", port=int(os.getenv("PORT", 5000)))
